import io
import logging
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ARGB hex colors (alpha=FF)
C_HEADER_BG  = 'FF232F3E'
C_HEADER_FG  = 'FFFFFFFF'
C_ORANGE_BG  = 'FFFF9900'
C_ORANGE_FG  = 'FF000000'
C_ALT_ROW    = 'FFF5F8FA'
C_WHITE      = 'FFFFFFFF'
C_GREEN_TAB  = 'FF1A7F37'
C_BLUE_TAB   = 'FF0073BB'

RESOURCE_COLUMNS = {
    'ec2_instances':          ['Name', 'Instance ID', 'Type', 'State', 'Private IP', 'Public IP', 'Region', 'VPC ID', 'Subnet ID', 'Key Name', 'Launch Time'],
    'vpcs':                   ['Name', 'VPC ID', 'CIDR', 'State', 'Default', 'Owner ID', 'Region'],
    'subnets':                ['Name', 'Subnet ID', 'VPC ID', 'CIDR', 'AZ', 'Available IPs', 'State', 'Map Public IP', 'Region'],
    'security_groups':        ['Name', 'Group ID', 'VPC ID', 'Description', 'Inbound Rules', 'Outbound Rules', 'Region'],
    'route_tables':           ['Name', 'Route Table ID', 'VPC ID', 'Main', 'Routes', 'Associations', 'Region'],
    'internet_gateways':      ['Name', 'IGW ID', 'VPC ID', 'State', 'Region'],
    'nat_gateways':           ['Name', 'NAT GW ID', 'VPC ID', 'Type', 'State', 'Subnet ID', 'Region'],
    'elastic_ips':            ['Name', 'Allocation ID', 'Public IP', 'Private IP', 'Instance ID', 'Domain', 'Region'],
    'volumes':                ['Name', 'Volume ID', 'Type', 'Size (GB)', 'State', 'AZ', 'Encrypted', 'Instance ID', 'Region'],
    'load_balancers_classic': ['Name', 'DNS Name', 'VPC ID', 'Scheme', 'Instances', 'Region'],
    'load_balancers_v2':      ['Name', 'ARN', 'DNS Name', 'VPC ID', 'Type', 'Scheme', 'State', 'Region'],
    'rds_instances':          ['Name', 'Engine', 'Instance Class', 'State', 'VPC ID', 'Multi-AZ', 'Storage (GB)', 'Endpoint', 'Region'],
    'rds_clusters':           ['Name', 'Engine', 'State', 'VPC ID', 'Multi-AZ', 'Members', 'Endpoint', 'Region'],
    'lambda_functions':       ['Name', 'Runtime', 'Handler', 'State', 'Memory (MB)', 'Timeout (s)', 'Code Size', 'VPC ID', 'Region'],
    'ecs_clusters':           ['Name', 'Status', 'Running Tasks', 'Pending Tasks', 'Active Services', 'Region'],
    'eks_clusters':           ['Name', 'Version', 'Status', 'VPC ID', 'Endpoint', 'Created', 'Region'],
    'elasticache_clusters':   ['Name', 'Engine', 'Node Type', 'Status', 'Nodes', 'VPC ID', 'Region'],
    'dynamodb_tables':        ['Name', 'Status', 'Items', 'Size (bytes)', 'Billing Mode', 'Region'],
    'sns_topics':             ['Name', 'ARN', 'Region'],
    'sqs_queues':             ['Name', 'URL', 'Region'],
    's3_buckets':             ['Name', 'Region', 'Created', 'Versioning'],
    'cloudfront_distributions': ['Name', 'ID', 'Domain', 'Status', 'Enabled', 'Price Class', 'Origins'],
    'route53_zones':          ['Name', 'ID', 'Private', 'Record Count', 'Comment'],
    'autoscaling_groups':     ['Name', 'Min', 'Max', 'Desired', 'Instances', 'Region'],
    'vpc_endpoints':          ['Name', 'Endpoint ID', 'VPC ID', 'Service', 'Type', 'State', 'Region'],
    'vpc_peering':            ['Name', 'Peering ID', 'State', 'Requester VPC', 'Requester CIDR', 'Requester Owner', 'Accepter VPC', 'Accepter CIDR', 'Accepter Owner', 'Region'],
    'transit_gateways':       ['Name', 'TGW ID', 'State', 'Description', 'Owner ID', 'Region'],
    'snapshots':              ['Name', 'Snapshot ID', 'Volume ID', 'State', 'Size (GB)', 'Encrypted', 'Start Time', 'Description', 'Region'],
    'opensearch_domains':     ['Name', 'ARN', 'Engine Version', 'Instance Type', 'Instances', 'State', 'Endpoint', 'VPC ID', 'Region'],
    'beanstalk_environments': ['Name', 'ID', 'Application', 'Solution Stack', 'Status', 'Health', 'Tier', 'URL', 'Region'],
    'acm_certificates':       ['Domain', 'ARN', 'Status', 'Type', 'In Use', 'Key Algorithm', 'Expires', 'Created', 'Region'],
    'iam_users':              ['Name', 'User ID', 'ARN', 'Path', 'Created', 'Password Last Used'],
    'iam_roles':              ['Name', 'Role ID', 'ARN', 'Path', 'Created', 'Description', 'Max Session (s)'],
    'iam_policies':           ['Name', 'Policy ID', 'ARN', 'Path', 'Created', 'Updated', 'Attachments'],
}

RESOURCE_FIELDS = {
    'ec2_instances':          ['name', 'id', 'type', 'state', 'private_ip', 'public_ip', 'region', 'vpc_id', 'subnet_id', 'key_name', 'launch_time'],
    'vpcs':                   ['name', 'id', 'cidr', 'state', 'is_default', 'owner_id', 'region'],
    'subnets':                ['name', 'id', 'vpc_id', 'cidr', 'az', 'available_ips', 'state', 'map_public_ip', 'region'],
    'security_groups':        ['name', 'id', 'vpc_id', 'description', 'inbound_rules', 'outbound_rules', 'region'],
    'route_tables':           ['name', 'id', 'vpc_id', 'main', 'routes', 'associations', 'region'],
    'internet_gateways':      ['name', 'id', 'vpc_id', 'state', 'region'],
    'nat_gateways':           ['name', 'id', 'vpc_id', 'type', 'state', 'subnet_id', 'region'],
    'elastic_ips':            ['name', 'id', 'public_ip', 'private_ip', 'instance_id', 'domain', 'region'],
    'volumes':                ['name', 'id', 'type', 'size', 'state', 'az', 'encrypted', 'instance_id', 'region'],
    'load_balancers_classic': ['name', 'dns_name', 'vpc_id', 'scheme', 'instances', 'region'],
    'load_balancers_v2':      ['name', 'id', 'dns_name', 'vpc_id', 'type', 'scheme', 'state', 'region'],
    'rds_instances':          ['name', 'engine', 'instance_class', 'state', 'vpc_id', 'multi_az', 'storage', 'endpoint', 'region'],
    'rds_clusters':           ['name', 'engine', 'state', 'vpc_id', 'multi_az', 'members', 'endpoint', 'region'],
    'lambda_functions':       ['name', 'runtime', 'handler', 'state', 'memory', 'timeout', 'code_size', 'vpc_id', 'region'],
    'ecs_clusters':           ['name', 'status', 'running_tasks', 'pending_tasks', 'active_services', 'region'],
    'eks_clusters':           ['name', 'version', 'status', 'vpc_id', 'endpoint', 'created', 'region'],
    'elasticache_clusters':   ['name', 'engine', 'node_type', 'status', 'nodes', 'vpc_id', 'region'],
    'dynamodb_tables':        ['name', 'status', 'items', 'size_bytes', 'billing_mode', 'region'],
    'sns_topics':             ['name', 'id', 'region'],
    'sqs_queues':             ['name', 'url', 'region'],
    's3_buckets':             ['name', 'region', 'created', 'versioning'],
    'cloudfront_distributions': ['name', 'id', 'domain', 'status', 'enabled', 'price_class', 'origins'],
    'route53_zones':          ['name', 'id', 'private', 'record_count', 'comment'],
    'autoscaling_groups':     ['name', 'min_size', 'max_size', 'desired', 'instances', 'region'],
    'vpc_endpoints':          ['name', 'id', 'vpc_id', 'service', 'type', 'state', 'region'],
    'vpc_peering':            ['name', 'id', 'state', 'requester_vpc', 'requester_cidr', 'requester_owner', 'accepter_vpc', 'accepter_cidr', 'accepter_owner', 'region'],
    'transit_gateways':       ['name', 'id', 'state', 'description', 'owner_id', 'region'],
    'snapshots':              ['name', 'id', 'volume_id', 'state', 'volume_size', 'encrypted', 'start_time', 'description', 'region'],
    'opensearch_domains':     ['name', 'id', 'engine_version', 'instance_type', 'instance_count', 'state', 'endpoint', 'vpc_id', 'region'],
    'beanstalk_environments': ['name', 'id', 'application', 'solution_stack', 'status', 'health', 'tier', 'url', 'region'],
    'acm_certificates':       ['name', 'id', 'status', 'type', 'in_use', 'key_algorithm', 'not_after', 'created', 'region'],
    'iam_users':              ['name', 'id', 'arn', 'path', 'created', 'password_last_used'],
    'iam_roles':              ['name', 'id', 'arn', 'path', 'created', 'description', 'max_session'],
    'iam_policies':           ['name', 'id', 'arn', 'path', 'created', 'updated', 'attachment_count'],
}

RESOURCE_LABELS = {
    'ec2_instances':          'EC2 Instances',
    'vpcs':                   'VPCs',
    'subnets':                'Subnets',
    'security_groups':        'Security Groups',
    'route_tables':           'Route Tables',
    'internet_gateways':      'Internet Gateways',
    'nat_gateways':           'NAT Gateways',
    'elastic_ips':            'Elastic IPs',
    'volumes':                'EBS Volumes',
    'load_balancers_classic': 'Classic Load Balancers',
    'load_balancers_v2':      'ALB/NLB Load Balancers',
    'rds_instances':          'RDS Instances',
    'rds_clusters':           'RDS Clusters',
    'lambda_functions':       'Lambda Functions',
    'ecs_clusters':           'ECS Clusters',
    'eks_clusters':           'EKS Clusters',
    'elasticache_clusters':   'ElastiCache Clusters',
    'dynamodb_tables':        'DynamoDB Tables',
    'sns_topics':             'SNS Topics',
    'sqs_queues':             'SQS Queues',
    's3_buckets':             'S3 Buckets',
    'cloudfront_distributions': 'CloudFront Distributions',
    'route53_zones':          'Route53 Hosted Zones',
    'autoscaling_groups':     'Auto Scaling Groups',
    'vpc_endpoints':          'VPC Endpoints',
    'vpc_peering':            'VPC Peering Connections',
    'transit_gateways':       'Transit Gateways',
    'snapshots':              'EBS Snapshots',
    'opensearch_domains':     'OpenSearch Domains',
    'beanstalk_environments': 'Beanstalk Environments',
    'acm_certificates':       'ACM Certificates',
    'iam_users':              'IAM Users',
    'iam_roles':              'IAM Roles',
    'iam_policies':           'IAM Policies',
}

# Resource types that are associated with a VPC
VPC_RESOURCE_TYPES = [
    'ec2_instances', 'subnets', 'security_groups', 'route_tables',
    'internet_gateways', 'nat_gateways', 'load_balancers_classic',
    'load_balancers_v2', 'rds_instances', 'rds_clusters', 'lambda_functions',
    'eks_clusters', 'elasticache_clusters', 'vpc_endpoints', 'vpc_peering',
    'opensearch_domains',
]

# Resource types that are global or not VPC-bound
GLOBAL_RESOURCE_TYPES = [
    's3_buckets', 'cloudfront_distributions', 'route53_zones',
    'dynamodb_tables', 'sns_topics', 'sqs_queues', 'transit_gateways',
    'autoscaling_groups', 'elastic_ips', 'volumes', 'ecs_clusters',
    'snapshots', 'beanstalk_environments', 'acm_certificates',
    'iam_users', 'iam_roles', 'iam_policies',
]


def _cell_val(v):
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'Yes' if v else 'No'
    return v


def _hdr(cell, bg, fg='FFFFFFFF', bold=True, wrap=False):
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    cell.font = Font(color=fg, bold=bold, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value or ''))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 55)


def _write_section(ws, row, rtype, resources):
    if not resources:
        return row

    columns = RESOURCE_COLUMNS.get(rtype, [])
    fields = RESOURCE_FIELDS.get(rtype, [])
    label = RESOURCE_LABELS.get(rtype, rtype)
    n_cols = len(columns) or 1

    # Section title row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    title_cell = ws.cell(row=row, column=1, value=f'  {label}  ({len(resources)})')
    _hdr(title_cell, C_ORANGE_BG, C_ORANGE_FG, bold=True)
    row += 1

    # Column header row
    for ci, col_name in enumerate(columns, 1):
        _hdr(ws.cell(row=row, column=ci, value=col_name), C_HEADER_BG, C_HEADER_FG)
    row += 1

    # Data rows
    for ri, resource in enumerate(resources):
        fill_color = C_ALT_ROW if ri % 2 else C_WHITE
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        for ci, field in enumerate(fields, 1):
            cell = ws.cell(row=row, column=ci, value=_cell_val(resource.get(field, '')))
            cell.fill = fill
            cell.alignment = Alignment(vertical='center')
        row += 1

    return row + 1  # blank spacer


def _sanitize_sheet_name(name, existing_names):
    for ch in r'\/?*[]':
        name = name.replace(ch, '-')
    name = name.replace(':', '-')[:31]
    base, counter = name, 1
    while name in existing_names:
        suffix = f'-{counter}'
        name = base[:31 - len(suffix)] + suffix
        counter += 1
    return name


def generate_excel(scan_results):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    vpcs = scan_results.get('vpcs', [])
    account_id = scan_results.get('_account_id', '')
    regions = scan_results.get('_regions', [])

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet('Summary')
    ws_sum.freeze_panes = 'A3'

    ws_sum.merge_cells('A1:D1')
    title = ws_sum.cell(row=1, column=1,
                        value=f'AWS Resource Inventory  |  Account: {account_id}  |  Regions: {len(regions)}')
    _hdr(title, C_HEADER_BG, C_HEADER_FG, bold=True)
    ws_sum.row_dimensions[1].height = 20

    for ci, hdr in enumerate(['Resource Type', 'Count', 'Has Resources', 'Regions'], 1):
        _hdr(ws_sum.cell(row=2, column=ci, value=hdr), C_ORANGE_BG, C_ORANGE_FG)

    for ri, (rtype, label) in enumerate(RESOURCE_LABELS.items(), 3):
        items = scan_results.get(rtype, [])
        region_set = sorted({r.get('region', '') for r in items if r.get('region')})
        fill_color = C_ALT_ROW if ri % 2 else C_WHITE
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        for ci, val in enumerate([label, len(items), 'Yes' if items else 'No', ', '.join(region_set)], 1):
            c = ws_sum.cell(row=ri, column=ci, value=val)
            c.fill = fill
            c.alignment = Alignment(vertical='center')

    _auto_width(ws_sum)

    # ── One sheet per VPC ────────────────────────────────────────────────────
    used_names = {'Summary'}
    region_colors = {}
    color_pool = [
        'FF0073BB', 'FF1A7F37', 'FF7B0E93', 'FFC7131F',
        'FFDF6E00', 'FF005276', 'FF5A3E00', 'FF00695C',
    ]
    color_idx = 0

    for vpc in sorted(vpcs, key=lambda v: (v.get('region', ''), v.get('name', '') or v.get('id', ''))):
        vpc_id = vpc['id']
        vpc_name = vpc.get('name', '') or vpc_id
        region = vpc.get('region', '')

        if region not in region_colors:
            region_colors[region] = color_pool[color_idx % len(color_pool)]
            color_idx += 1

        raw_name = f"{region[:8]}-{vpc_name[:20]}"
        sheet_name = _sanitize_sheet_name(raw_name, used_names)
        used_names.add(sheet_name)

        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = region_colors[region][2:]  # strip leading 'FF'
        ws.freeze_panes = 'A3'

        ws.merge_cells('A1:K1')
        info = ws.cell(
            row=1, column=1,
            value=f'VPC: {vpc_name}  ({vpc_id})  |  Region: {region}  |  CIDR: {vpc.get("cidr", "")}  |  Default: {"Yes" if vpc.get("is_default") else "No"}'
        )
        _hdr(info, C_HEADER_BG, C_HEADER_FG)
        ws.row_dimensions[1].height = 18

        row = 3
        for rtype in VPC_RESOURCE_TYPES:
            all_res = scan_results.get(rtype, [])
            vpc_res = [r for r in all_res if r.get('vpc_id') == vpc_id]
            row = _write_section(ws, row, rtype, vpc_res)

        _auto_width(ws)

    # ── Global / non-VPC resources ───────────────────────────────────────────
    ws_global = wb.create_sheet('Global Resources')
    ws_global.sheet_properties.tabColor = C_BLUE_TAB[2:]
    ws_global.freeze_panes = 'A3'

    ws_global.merge_cells('A1:K1')
    _hdr(ws_global.cell(row=1, column=1, value='Global & Non-VPC Resources'), C_HEADER_BG, C_HEADER_FG)
    used_names.add('Global Resources')

    row = 3
    for rtype in GLOBAL_RESOURCE_TYPES:
        resources = scan_results.get(rtype, [])
        row = _write_section(ws_global, row, rtype, resources)

    _auto_width(ws_global)

    # ── Unattached / no-VPC resources (VPC-type but no vpc_id) ──────────────
    ws_un = wb.create_sheet('Unattached')
    ws_un.freeze_panes = 'A3'

    ws_un.merge_cells('A1:K1')
    _hdr(ws_un.cell(row=1, column=1, value='VPC-Scoped Resources With No VPC Association'),
         C_HEADER_BG, C_HEADER_FG)

    row = 3
    for rtype in VPC_RESOURCE_TYPES:
        unattached = [r for r in scan_results.get(rtype, []) if not r.get('vpc_id')]
        row = _write_section(ws_un, row, rtype, unattached)

    _auto_width(ws_un)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
